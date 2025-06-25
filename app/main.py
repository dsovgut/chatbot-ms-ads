# app/main.py

# --- Imports ---
import chainlit as cl
from openai import AsyncOpenAI
import os
from dotenv import load_dotenv
import fitz  # PyMuPDF for PDFs
import docx  # python-docx for DOCX
import openpyxl # for XLSX
import csv      # built-in for CSV
import json     # built-in for JSON
import io       # built-in for file handling
import asyncio # Added for potential async operations in RAG
import sys

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from rag import RAG

# --- Load Environment Variables ---
load_dotenv()

# --- Initialize OpenAI Client ---
# This might be used by your RAG system as well, or the RAG system might have its own LLM client.
# Your teammate will clarify if this specific client instance is needed for her RAG part.
client = AsyncOpenAI(api_key=os.getenv("OPENAI_API_KEY"))
cl.instrument_openai() # This is for Chainlit's OpenAI integration, may or may not be relevant if RAG handles its own LLM.

# --- System Prompt for LLM ---
# ### Teammate Integration Point ###
# This SYSTEM_PROMPT will likely be used by the LLM *within* the RAG system.
# Or, if the RAG system only retrieves context and this frontend still makes the final LLM call,
# then this prompt will be used here. Coordinate with your teammate.
SYSTEM_PROMPT = """You are an expert AI assistant for the University of Chicago's MS in Applied Data Science program.
Your goal is to provide accurate and helpful information to prospective students, current enrollees, and alumni.
Use the context retrieved from the official program knowledge base to answer questions.
If the retrieved context provides the answer, base your response primarily on that information and clearly state that the information comes from the knowledge base.
If the retrieved context does not fully answer the question, you can use your general knowledge but clearly indicate that the specific detail was not found in the knowledge base.
Maintain a professional, helpful, and concise tone. Be mindful of responsible AI practices and avoid making up information not supported by the provided context or your training."""

# --- File Processing Functions ---
# (These functions can remain as they are. They might be useful if your RAG system
# can also incorporate context from user-uploaded files alongside its main knowledge base.)

def process_pdf(file_path: str) -> str:
    text = ""
    try:
        doc = fitz.open(file_path)
        for page in doc:
            page_text = page.get_text("text")
            if page_text:
                text += page_text + "\n\n"
        doc.close()
    except Exception as e:
        print(f"Error processing PDF {file_path}: {e}")
        raise
    return text.strip()

def process_docx(file_path: str) -> str:
    text = ""
    try:
        document = docx.Document(file_path)
        for para in document.paragraphs:
            text += para.text + "\n\n"
    except Exception as e:
        print(f"Error processing DOCX {file_path}: {e}")
        raise
    return text.strip()

def process_txt(file_path: str) -> str:
    text = ""
    try:
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            text = f.read()
    except Exception as e:
        print(f"Error processing TXT {file_path}: {e}")
        raise
    return text.strip()

def process_csv(file_path: str) -> str:
    text = ""
    try:
        with open(file_path, "r", encoding="utf-8", errors="ignore", newline='') as f:
            reader = csv.reader(f)
            for row in reader:
                text += ", ".join(cell.strip() for cell in row) + "\n"
    except Exception as e:
        print(f"Error processing CSV {file_path}: {e}")
        raise
    return text.strip()

def process_json(file_path: str) -> str:
    text = ""
    try:
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            data = json.load(f)
            text = json.dumps(data, indent=2)
    except Exception as e:
        print(f"Error processing JSON {file_path}: {e}")
        raise
    return text.strip()

def process_xlsx(file_path: str) -> str:
    text = ""
    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        for sheet_name in workbook.sheetnames:
            text += f"--- Sheet: {sheet_name} ---\n"
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                row_texts = []
                for cell in row:
                    if cell.value is not None:
                        row_texts.append(str(cell.value).strip())
                if row_texts:
                    text += ", ".join(row_texts) + "\n"
            text += "\n"
    except Exception as e:
        print(f"Error processing XLSX {file_path}: {e}")
        raise
    return text.strip()

# --- File Handler Dispatch Dictionary ---
FILE_HANDLERS = {
    "application/pdf": process_pdf,
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": process_docx,
    ".docx": process_docx,
    "text/plain": process_txt,
    ".txt": process_txt,
    "text/csv": process_csv,
    ".csv": process_csv,
    "application/json": process_json,
    ".json": process_json,
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": process_xlsx,
    ".xlsx": process_xlsx,
}

# --- Placeholder for RAG System Call ---

# Instantiate RAG system
rag_system = RAG()
async def get_rag_response(user_query: str, chat_history: list) -> str:
    """
    RAG System call to return response to user query.
    """
    print(f"--- RAG SYSTEM CALL (Placeholder) ---")
    print(f"Query: {user_query}")

    try:
        response = await rag_system.query(user_query)
        return response
    except Exception as e:
        print(f"Error during RAG query: {e}")
        return "Sorry, I couldn't process your request at this time."

# --- Chainlit Event Handlers ---

@cl.on_chat_start
async def start_chat():
    """Initializes session history."""
    cl.user_session.set("history", [{"role": "system", "content": SYSTEM_PROMPT}]) # Preload history with system prompt
    print("Chat session started, history initialized with system prompt.")

@cl.on_message
async def handle_message(msg: cl.Message):
    """Handles messages, calls RAG system (or LLM as fallback), and streams response."""
    history = cl.user_session.get("history", []) # Includes system prompt now
    user_query = ""
    if msg.content: # Ensure msg.content is not None
        user_query = msg.content
        history.append({"role": "user", "content": user_query})

    # --- Process Attached Files (kept for potential use) ---
    extracted_texts_for_rag = [] # Contexts from uploaded files
    # (Your existing file processing logic can go here if you intend for the RAG system
    # to optionally use context from user-uploaded files. If not, you can simplify or remove this part.)
    if msg.elements:
        # ... (your existing file processing loop to populate extracted_texts_for_rag) ...
        # For brevity, I'm omitting the full file processing loop here, but you'd copy it
        # if you want to pass `extracted_texts_for_rag` to `get_rag_response`.
        # Make sure it populates `extracted_texts_for_rag` similar to how `extracted_texts` was populated.
        print(f"--- DEBUG: {len(msg.elements)} file(s) attached by user. Processing... ---")
        # This part needs to be re-integrated carefully if you want to use uploaded files with RAG.
        # For simplicity in this example, let's assume for now RAG primarily uses its own knowledge base.
        # Your teammate can advise on how best to pass extra context from uploaded files to her RAG system.


    # ### Teammate Integration Point ###
    # Call the RAG system to get the response
    try:
        # Pass the user_query, the conversational history, and any context from uploaded files
        final_answer = await get_rag_response(user_query, history)
        
        response_msg = cl.Message(content="")
        await response_msg.send() # Send an empty message first to get the UI element
        await response_msg.stream_token(final_answer) # Stream the final answer
        await response_msg.update() # Finalize

        if final_answer: # Save RAG response to history
            history.append({"role": "assistant", "content": final_answer})

    except Exception as e:
        error_message = f"Sorry, I encountered an error getting a response: {e}"
        await cl.Message(content=error_message).send()
        print(f"Error in RAG system call or displaying response: {e}")

    cl.user_session.set("history", history) # Update history